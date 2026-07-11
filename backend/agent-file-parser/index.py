import json
import os
import base64
import io
import psycopg2
import psycopg2.extras

SCHEMA = "t_p84565078_code_expression_proj"
ENERGY_FILE = 30

def tbl(name):
    return f"{SCHEMA}.{name}"

def get_db():
    return psycopg2.connect(os.environ["DATABASE_URL"])

CORS = {
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Methods": "POST, OPTIONS",
    "Access-Control-Allow-Headers": "Content-Type, X-Session-Id",
}

def ok(data):
    return {"statusCode": 200, "headers": CORS, "body": json.dumps(data, ensure_ascii=False, default=str)}

def err(msg, code=400):
    return {"statusCode": code, "headers": CORS, "body": json.dumps({"error": msg}, ensure_ascii=False)}

def get_session_user(event, conn):
    session_id = (event.get("headers") or {}).get("X-Session-Id", "")
    if not session_id:
        return None
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        f"SELECT u.* FROM {tbl('lk_sessions')} s "
        f"JOIN {tbl('lk_users')} u ON u.id = s.user_id "
        f"WHERE s.id = %s AND s.expires_at > NOW() AND u.is_active = TRUE",
        (session_id,)
    )
    return cur.fetchone()

def get_free_used(conn, user_id: int) -> int:
    cur = conn.cursor()
    cur.execute(
        f"SELECT free_used FROM {tbl('salon_agent_free_usage')} WHERE user_id = %s",
        (user_id,)
    )
    row = cur.fetchone()
    return row[0] if row else 0

def get_salon_balance(conn, salon_id: int) -> int:
    cur = conn.cursor()
    cur.execute(f"SELECT credits_balance FROM {tbl('salons')} WHERE id = %s", (salon_id,))
    row = cur.fetchone()
    return row[0] if row else 0

def deduct_energy(conn, salon_id: int, user_id: int, amount: int, action: str):
    cur = conn.cursor()
    cur.execute(
        f"UPDATE {tbl('salons')} SET credits_balance = credits_balance - %s WHERE id = %s",
        (amount, salon_id)
    )
    cur.execute(
        f"INSERT INTO {tbl('credit_transactions')} (salon_id, user_id, action, amount, tool_key, type) "
        f"VALUES (%s, %s, %s, %s, %s, 'debit')",
        (salon_id, user_id, action, amount, "agent_file")
    )

def parse_csv(data: bytes) -> str:
    import csv
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return ""
    lines = []
    for row in rows[:200]:
        lines.append(" | ".join(cell.strip() for cell in row))
    if len(rows) > 200:
        lines.append(f"... (показано 200 из {len(rows)} строк)")
    return "\n".join(lines)

def parse_excel(data: bytes) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    result = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        result.append(f"=== Лист: {sheet_name} ===")
        rows_written = 0
        for row in ws.iter_rows(values_only=True):
            if all(cell is None for cell in row):
                continue
            result.append(" | ".join(str(cell) if cell is not None else "" for cell in row))
            rows_written += 1
            if rows_written >= 200:
                result.append("... (показано 200 строк)")
                break
    return "\n".join(result)

def parse_pdf(data: bytes) -> str:
    import pdfplumber
    text_parts = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for i, page in enumerate(pdf.pages[:20]):
            text = page.extract_text()
            if text:
                text_parts.append(text.strip())
            if i >= 19:
                text_parts.append("... (показано первые 20 страниц)")
    return "\n\n".join(text_parts)

def parse_docx(data: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(data))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs[:500])

def handler(event: dict, context) -> dict:
    """Парсинг файлов для бизнес-ассистента (PDF, Excel, CSV, DOCX). Списывает 30 энергии."""
    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": CORS, "body": ""}

    if event.get("httpMethod") != "POST":
        return err("Method not allowed", 405)

    conn = get_db()
    try:
        user = get_session_user(event, conn)
        if not user:
            return err("Unauthorized", 401)

        if user["role"] not in ("owner", "admin", "solo_master") and not user.get("is_admin"):
            return err("Forbidden", 403)

        free_used = get_free_used(conn, user["id"])
        is_free_mode = free_used < 10

        if is_free_mode:
            return err("file_not_available_in_free_mode", 403)

        salon_id = user.get("salon_id")
        if not salon_id:
            return err("no_salon", 400)

        balance = get_salon_balance(conn, salon_id)
        if balance < ENERGY_FILE:
            return err("no_energy", 402)

        body = json.loads(event.get("body") or "{}")
        file_b64 = body.get("file_base64", "")
        filename = body.get("filename", "file")
        if not file_b64:
            return err("no_file", 400)

        file_data = base64.b64decode(file_b64)
        name_lower = filename.lower()

        if name_lower.endswith(".csv"):
            extracted = parse_csv(file_data)
            file_type = "CSV-таблица"
        elif name_lower.endswith(".xlsx") or name_lower.endswith(".xls"):
            extracted = parse_excel(file_data)
            file_type = "Excel-таблица"
        elif name_lower.endswith(".pdf"):
            extracted = parse_pdf(file_data)
            file_type = "PDF-документ"
        elif name_lower.endswith(".docx") or name_lower.endswith(".doc"):
            extracted = parse_docx(file_data)
            file_type = "Word-документ"
        else:
            return err("unsupported_format", 400)

        if not extracted or not extracted.strip():
            return err("empty_file", 400)

        deduct_energy(conn, salon_id, user["id"], ENERGY_FILE, "file_upload")
        conn.commit()

        balance_after = get_salon_balance(conn, salon_id)

        return ok({
            "extracted_text": extracted,
            "file_type": file_type,
            "filename": filename,
            "energy_spent": ENERGY_FILE,
            "energy_balance": balance_after,
        })

    finally:
        conn.close()